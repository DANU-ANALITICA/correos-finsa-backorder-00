# main.py
## librerías
import os
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import imaplib
import time 
from dotenv import load_dotenv
import pandas as pd
import re
import numpy as np
import ast
from google.cloud import bigquery
from email_validator import validate_email, EmailNotValidError
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

EJECUTION_MODE = "PRUEBA"

## Extrae variables del .env
load_dotenv()
PROJECT_ID = os.environ["GCP_PROJECT"]
BUYERS_PASSWORD = os.environ.get("BUYERS_PASSWORD", "")
BUYERS_PASSWORD = ast.literal_eval(BUYERS_PASSWORD)
buyer_passwords = pd.DataFrame.from_dict(BUYERS_PASSWORD, orient='index', columns=['Password'])
buyer_passwords.index.name = 'Usuario_ID'
GOOGLE_CREDENTIALS = os.environ["GOOGLE_APPLICATION_CREDENTIALS"]

client = bigquery.Client(project=PROJECT_ID)

DATASET_Y_TABLA = "raw_data.historico_correos_backorder"  
TABLA_COMPLETA_ID = f"{PROJECT_ID}.{DATASET_Y_TABLA}"

job_config = bigquery.LoadJobConfig(
    write_disposition="WRITE_APPEND" 
)

query = '''SELECT DISTINCT 
p.Proveedor as Proveedor, 
p.Nombre as Nombre,
p.Email1 AS Email,
p.Email2 AS Email_CC,
p.Comprador as Comprador,
u.email as Email_COMPRADOR
FROM `finsadashboard.raw_data.Proveedores` as p
LEFT JOIN `finsadashboard.raw_data.Compradores` as c
ON p.Comprador = c.Comprador
LEFT JOIN `finsadashboard.raw_data.Usuarios` as u
ON c.Usuario = u.Usuario
WHERE 
COALESCE(
  NULLIF(p.Email1, ''),
  NULLIF(p.Email2, ''),
  NULLIF(p.Email3, '')
) LIKE '%@%'  
AND p.Comprador IS NOT NULL 
AND p.Proveedor IN  (Select DISTINCT Proveedor from `finsadashboard.mrts.mrts_backorder_MTY`)
AND NULLIF(u.email, '') LIKE '%@%'  
'''


# # Query de prueba para envío a destinatario fijo
# query = '''SELECT DISTINCT Proveedor, Nombre,
#     Case When Proveedor = 95 Then 'keyla.islas@danuanalitica.com' Else 'daniel.perez@danuanalitica.com' End AS Email,
#     case when Proveedor = 95 then 'daniel.perez@danuanalitica.com' else 'keyla.islas@danuanalitica.com' end AS Email_CC,
#     7.0 AS Comprador
# FROM `finsadashboard.raw_data.Proveedores` 
# --LIMIT 3
# WHERE Proveedor IN (95,21)
# '''

correos = client.query(query).to_dataframe()


query = """
        SELECT 
            bo.EDP,
            bo.FOLIO AS ORDEN_COMPRA,
            bo.PARTIDA,
            bo.FECHA_ALTA,
            bo.ARTICULO,
            bo.DESCRIPCION,
            bo.CANTIDAD,
            bo.CANTIDAD_RECIBIDA,
            bo.BACKORDER,
            bo.UNIDAD,
            bo.COSTO,
            bo.FECHA_EMBARQUE,
            bo.DIAS_RETRASO_EMBARQUE,
            bo.NOMBRE_SUCURSAL,
            bo.NOMBRE_ALMACEN,
            bo.NOMBRE_COMPRADOR,
            bo.NOMBRE_PROVEEDOR,
            bo.PROVEEDOR,
            bo.COMPRADOR,
            bo.SUCURSAL, 
            bo.ALMACEN,
            bo.PEDIDO,
            --check.Envio_Backorder AS Envio_Backorder
        FROM `finsadashboard.mrts.mrts_backorder_MTY` bo
        LEFT JOIN `finsadashboard.mrts.checkbox_state_proveedores` AS check
        ON CONCAT(bo.PROVEEDOR, '|', bo.SUCURSAL) = check.ID
        where Envio_Backorder = True
        ORDER BY NOMBRE_PROVEEDOR, FECHA_ALTA
    """

backorder = client.query(query).to_dataframe()
print(backorder)

loop_values = backorder[['NOMBRE_COMPRADOR','NOMBRE_PROVEEDOR', 'NOMBRE_SUCURSAL', 'NOMBRE_ALMACEN', 'PROVEEDOR', 'COMPRADOR', 'SUCURSAL','ALMACEN']].drop_duplicates()
loop_values = loop_values.merge(correos, left_on='PROVEEDOR', right_on='Proveedor', how='left')
loop_values = loop_values.merge(buyer_passwords, left_on='COMPRADOR', right_index=True, how='left')
loop_values = loop_values[['NOMBRE_PROVEEDOR','NOMBRE_COMPRADOR', 'NOMBRE_SUCURSAL', 'NOMBRE_ALMACEN','PROVEEDOR', 'COMPRADOR', 'SUCURSAL','ALMACEN','Email_COMPRADOR', 'Email','Email_CC', 'Password']]


##─── DEF CORREOS_CLEAN ─────────────────────────────────────────────────────────────────

def clean_email_addresses(correos: pd.DataFrame) -> pd.DataFrame:
    extraido = correos['Email'].str.extract(r'\[(.*?)\]', expand=False)
    correos['Clean'] = extraido.fillna(correos['Email'])
    extraido = correos['Clean'].str.extract(r'<(.*?)>', expand=False)
    correos['Clean'] = (
    extraido
    .fillna(correos['Clean'])
    .str.replace(r'[\[\]<>]', '', regex=True)  
    .str.rsplit(' ', n=1).str[-1]              
    .str.rsplit(':', n=1).str[-1]           
)
    patron = r'([a-zA-Z0-9Ññ._-]+@[a-zA-Z0-9Ññ_-]+(?:\.[a-zA-Z]{2,})+)'    
    correos['Clean'] = correos['Clean'].str.extract(patron)
    correos = correos.dropna(subset=['Clean'])

    extraido = correos['Email_CC'].str.extract(r'\[(.*?)\]', expand=False)
    correos['Clean_CC'] = extraido.fillna(correos['Email_CC'])
    extraido = correos['Clean_CC'].str.extract(r'<(.*?)>', expand=False)
    correos['Clean_CC'] = (
    extraido
    .fillna(correos['Clean_CC'])
    .str.replace(r'[\[\]<>]', '', regex=True)  
    .str.rsplit(' ', n=1).str[-1]              
    .str.rsplit(':', n=1).str[-1]           
)
    patron = r'([a-zA-Z0-9Ññ._-]+@[a-zA-Z0-9Ññ_-]+(?:\.[a-zA-Z]{2,})+)'    
    correos['Clean_CC'] = correos['Clean_CC'].str.extract(patron)
    correos = correos.dropna(subset=['Clean_CC'])

    extraido = correos['Email_COMPRADOR'].str.extract(r'\[(.*?)\]', expand=False)
    correos['Clean_COMPRADOR'] = extraido.fillna(correos['Email_COMPRADOR'])
    extraido = correos['Clean_COMPRADOR'].str.extract(r'<(.*?)>', expand=False)
    correos['Clean_COMPRADOR'] = (
    extraido
    .fillna(correos['Clean_COMPRADOR'])
    .str.replace(r'[\[\]<>]', '', regex=True)  
    .str.rsplit(' ', n=1).str[-1]              
    .str.rsplit(':', n=1).str[-1]           
)
    patron = r'([a-zA-Z0-9Ññ._-]+@[a-zA-Z0-9Ññ_-]+(?:\.[a-zA-Z]{2,})+)'    
    correos['Clean_COMPRADOR'] = correos['Clean_COMPRADOR'].str.extract(patron)
    correos = correos.dropna(subset=['Clean_COMPRADOR'])
    
    return correos

##─── DEF VALIDAR DOMINIO ─────────────────────────────────────────────────────────────────
def verificar_existencia_correo(correo):
    try:
        validate_email(correo, check_deliverability=True)
        return True
    except EmailNotValidError as e:
        return False


##─── DEF DATA ─────────────────────────────────────────────────────────────────
def get_backorder(provider: int, branch: int, buyer: float, df: pd.DataFrame) -> pd.DataFrame:

    df = df[(df['PROVEEDOR'] == provider) &
            (df['SUCURSAL'] == branch) &
            (df['COMPRADOR'] == buyer)]
    df = df[['FECHA_ALTA','ORDEN_COMPRA','PEDIDO','ARTICULO','DESCRIPCION','EDP','PARTIDA','CANTIDAD','CANTIDAD_RECIBIDA','BACKORDER','UNIDAD','FECHA_EMBARQUE','DIAS_RETRASO_EMBARQUE']]
    df.columns = ['FECHA ALTA','ORDEN DE COMPRA','PEDIDO','ARTICULO','DESCRIPCIÓN','EDP','PARTIDA','CANTIDAD','CANTIDAD RECIBIDA','BACKORDER','UNIDAD','FECHA DE EMBARQUE','DIAS DE RETRASOEMBARQUE']
    return df

def limpiar_cadena(cadena: str) -> str:
    if pd.isna(cadena):
        return ""
    ##Eliminar caracteres especiales y acentos
    cadena = re.sub(r'[^a-zA-Z0-9]', '', cadena)
    return cadena.strip()

##─── CORREO ───────────────────────────────────────────────────────────────────
def send_email_backorder(df: pd.DataFrame, Proveedor : str, Email_proveedor: str, Email_comprador: str, Comprador: str,  Sucursal: str, Password: str, Almacen: str, Email_CC: str) -> None:
    if df.empty:
        print("El reporte está vacío, no se envía correo.")
        return

    html = f"""
    <html>
    <head>
    <style>
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            color: #333333;
            margin: 20px;
        }}
        h2 {{
            color: #0d2c54;
            border-bottom: 2px solid #0d2c54;
            padding-bottom: 8px;
        }}
        table {{
            border-collapse: collapse;
            width: 100%;
            margin-top: 15px;
            font-size: 13px;
        }}
        th {{
            background-color: #0d2c54;
            color: white;
            padding: 10px 12px;
            text-align: left;
            font-weight: 600;
        }}
        td {{
            padding: 10px 12px;
            border-bottom: 1px solid #dddddd;
        }}
        tr:nth-child(even) {{
            background-color: #f8f9fa;
        }}
        tr:hover {{
            background-color: #f1f3f5;
        }}
    </style>
    </head>
    <body>
    <p><b>Buen día estimado proveedor espero y se encuentre bien</b></p>
    <p><b>Anexo al presente un archivo en Excel que contiene todas las líneas abiertas pendientes de entrega (BackOrder), por favor le solicitamos enviar las fechas de entrega de cada una de las partidas en la mayor brevedad posible, sabiendo que las fechas indicadas son de arribo en almacén de FINSA.</b></p>
    <p><b>Para Finsa y proveedores es importante mantener al cliente final informado sobre la entrega de sus productos</b></p>
    <p><b>Agradecemos su puntal apoyo</b></p>
    <p><b>Saludos</b></p>
    <h2>Reporte de Backorder - {Sucursal}</h2>
    {df.to_html(index=False)}
    <br>
    <p style="font-size: 11px; color: #777777;">Este es un correo automático generado por el sistema.</p>
    </body>
    </html>
    """

    prov = limpiar_cadena(Proveedor)
    suc = limpiar_cadena(Sucursal)

    excel_file = f"Backorder_{prov}_{suc}.xlsx"

    df1 = df.copy()
    df1['NUEVA FECHA COMPROMISO'] = ""

    with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
        df1.to_excel(writer, index=False, sheet_name='Datos')
        
        ws = writer.sheets['Datos']
        
        header_fill = PatternFill("solid", fgColor="4472C4")
        for cell in ws[1]:  # fila 1 = encabezados
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 4

    msg = MIMEMultipart("mixed") #alternative
    msg["Subject"] = f"Reporte de Backorder - {Proveedor}"
    msg["From"]    = Email_comprador
    msg["To"]      = Email_proveedor #", ".join([r.strip() for r in RECEPTOR.split(",")])

    if Email_CC:
        CCs = Email_comprador + "," + Email_CC
    else:
        CCs = Email_comprador

    msg["Cc"]      = CCs ## "ruben.garza@finsa.com.mx" + "," +
    msg['User-Agent'] = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) Python-SMTPLIB"
    msg['X-Mailer'] = "Python-SMTP-Client"
    msg.attach(MIMEText(html, "html"))

    try:
        with open(excel_file, "rb") as f:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(f.read())
        
        encoders.encode_base64(part)
        
        part.add_header(
            "Content-Disposition",
            f"attachment; filename={excel_file}"
        )
        msg.attach(part)
        print("📦 Archivo Excel empaquetado correctamente")

    except FileNotFoundError:
        print(f"❌ Error: No se encontró el archivo {excel_file}")
        return

    if not Email_comprador or not Password:
        print("⚠️ Advertencia: Email_comprador o Password no están configurados en el archivo .env. No se puede enviar el correo.")
        return
    
    HOST =  "finsa--com--mx.criticalmail.net"

    try:
        with smtplib.SMTP_SSL(HOST, 465, timeout=30) as smtp:
            time.sleep(1)
            smtp.login(Email_comprador, Password)
            smtp.send_message(msg)
            print("✅ Correo enviado")

            if EJECUTION_MODE != "PRUEBA":

                try:
                    print(f"Subiendo datos a {TABLA_COMPLETA_ID}...")
                    df = df.copy()
                    df['FECHA_ENVIO'] = pd.Timestamp.now()
                    df['Email_proveedor'] = Email_proveedor
                    df['Email_comprador'] = Email_comprador
                    df['Comprador'] = Comprador
                    df['Sucursal'] = Sucursal
                    df['Proveedor'] = Proveedor
                    df['Almacen'] = Almacen   
                    job = client.load_table_from_dataframe(df, TABLA_COMPLETA_ID, job_config=job_config)
                    job.result()
                        
                    print(f"¡Tabla subida exitosamente! Se cargaron {job.output_rows} filas.")

                except Exception as e:
                    print(f"Error al subir los datos a BigQuery: {e}")
        
    except Exception as e:
        print("❌ Error al enviar el correo:", e)

    try:
        with imaplib.IMAP4_SSL(HOST, 993, timeout=30) as imap:
            imap.login(Email_comprador, Password)
            carpeta_enviados = "Sent" 
            
            imap.append(
                carpeta_enviados, 
                r'\Seen', 
                imaplib.Time2Internaldate(time.time()), 
                msg.as_bytes()
            )
            print("Copia guardada con éxito en la carpeta de Enviados (IMAP).")
    except Exception as e:
        print(f"No se pudo guardar la copia en Enviados: {e}")
        

## PROVEEDOR X PROVEEDOR

if __name__ == "__main__":
    try:
        loop_values = clean_email_addresses(loop_values)
        resultados = loop_values['Clean'].apply(verificar_existencia_correo)
        loop_values['es_valido'] = resultados
        resultados = loop_values['Clean_CC'].apply(verificar_existencia_correo)
        loop_values['es_valido_CC'] = resultados
        resultados = loop_values['Clean_COMPRADOR'].apply(verificar_existencia_correo)
        loop_values['es_valido_COMPRADOR'] = resultados
        loop_values = loop_values[loop_values['es_valido'] == True]
        loop_values = loop_values[loop_values['es_valido_COMPRADOR'] == True]
        loop_values['Clean_CC'] = np.where(~loop_values['es_valido_CC'], loop_values['Clean_CC'], None)

        loop_values.to_excel("nombre_archivo.xlsx", index=False)
      

        for row in loop_values.itertuples():
            print(row.NOMBRE_PROVEEDOR, row.Email)
            print("Obteniendo datos de backorder",row.NOMBRE_PROVEEDOR)
            df = get_backorder(row.PROVEEDOR, row.SUCURSAL, row.COMPRADOR, backorder)
            print(df)
            print("\nEnviando reporte por correo...")
            send_email_backorder(df, row.NOMBRE_PROVEEDOR, row.Email, row.Email_COMPRADOR, row.NOMBRE_COMPRADOR, row.NOMBRE_SUCURSAL, row.Password, row.NOMBRE_ALMACEN, row.Email_CC)

    except Exception as e:
        print("Ocurrió un error al consultar BigQuery:", e)